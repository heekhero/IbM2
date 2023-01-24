from openpyxl import load_workbook
import os

import numpy as np

from openpyxl.styles import PatternFill

def acc_reformat(acc):
    if isinstance(acc, str):
        true_acc = float(acc.split('+')[0].strip())
    elif isinstance(acc, float):
        true_acc = acc
    else:
        raise None
    rt_acc = round(true_acc, 1)
    return rt_acc

def get_row(ws, row_id):
    rt_list = []
    for col in ws[row_id]:
        col_v = col.value
        if col_v is None:
            col_v = ''
        rt_list.append(str(col_v))
    return rt_list

wb = load_workbook('/mnt/data3/fumh/FSL/BSearch/RealFewShot_new_no_cyan/results/tmp_in_data3_latest.xlsx')
ws = wb['compare_search_continue_cw']


th_idx = 11
_1shot_idx = [22, 23]
_2shot_idx = [26, 27]
_3shot_idx = [30, 31]
_4shot_idx = [33, 34]
_5shot_idx = [36, 37]
_8shot_idx = [40, 41]
_16shot_idx = [43, 44]
_1pt_idx = [46, 47]
_10pt_idx = [50, 51]

shot_col_dict = {'1' : _1shot_idx,
                 '2' : _2shot_idx,
                 '3' : _3shot_idx,
                 '4' : _4shot_idx,
                 '5' : _5shot_idx,
                 '8' : _8shot_idx,
                 '16' : _16shot_idx,
                 '1pt' : _1pt_idx,
                 '10pt' : _10pt_idx}


a = [3, 4, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19, 21, 22, 24, 25, 27, 28, 30, 31, 37, 38, 40, 41, 43, 44, 46, 47, 49, 50, 52, 53, 55, 56, 58, 59, 61, 62, 64, 65]
for i in a:
    for shot in ['1', '2', '3', '4', '5', '8', '16', '1pt', '10pt']:
        naive_acc_before = ws.cell(i, shot_col_dict[shot][0]+1).value
        bsearch_acc_before = ws.cell(i, shot_col_dict[shot][1]+1).value

        try:
            naive_acc_after = acc_reformat(naive_acc_before)
            bsearch_acc_after = acc_reformat(bsearch_acc_before)
        except:
            print(i, shot)
            continue

        naive_acc_write = str(naive_acc_after)
        gain = round(bsearch_acc_after - naive_acc_after, 1)
        gain_str = str(gain)
        if gain > -1e-6:
            gain_str = '+' + gain_str
        bsearch_acc_write = str(bsearch_acc_after) + " " +  gain_str

        if gain > 0.5:
            fille = PatternFill('solid', fgColor='228b22')
        elif gain > -1e-6:
            fille = PatternFill('solid', fgColor='00ff00')
        else:
            fille = PatternFill('solid', fgColor='ff0000')

        ws.cell(i, shot_col_dict[shot][0] + 1).value = naive_acc_write
        ws.cell(i, shot_col_dict[shot][1] + 1).fill = fille
        ws.cell(i, shot_col_dict[shot][1] + 1).value = bsearch_acc_write

wb.save('/mnt/data3/fumh/FSL/BSearch/RealFewShot_new_no_cyan/results/tmp_in_data3_latest.xlsx')