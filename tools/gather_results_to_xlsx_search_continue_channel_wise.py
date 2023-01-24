from openpyxl import load_workbook
import os
import sys

import init_path
from utils import max_acc_check



def acc_reformat(acc):
    return acc.split('+-')[0] + ' + ' + acc.split('+-')[1]

def get_row(ws, row_id):
    rt_list = []
    for col in ws[row_id]:
        col_v = col.value
        if col_v is None:
            col_v = ''
        rt_list.append(str(col_v))
    return rt_list

def check_location_charact(row, rule_info):
    for rule, rule_idx in rule_info:
        for jdx in rule_idx:
            if row[jdx] != rule:
                return False
    return True

wb = load_workbook('/mnt/data3/fumh/FSL/BSearch/RealFewShot_new_no_cyan/results/tmp_in_data3_latest.xlsx')
ws = wb['channel_wise_continue_search']

th_idx = 11
_1shot_idx = [22, 23, 24]
_2shot_idx = [27, 28, 29]
_3shot_idx = [32, 33, 34]
_4shot_idx = [36, 37, 38]
_5shot_idx = [40, 41, 42]
_8shot_idx = [45, 46, 47]
_16shot_idx = [49, 50, 51]
_1pt_idx = [53, 54, 55]
_10pt_idx = [58, 59, 60]

pre_idx = [12,]
cycle_idx = [13,]
final_idx = [14, 18]
batch_size_idx = [7, 17]

shot_col_dict = {'1shot' : _1shot_idx,
                 '2shot' : _2shot_idx,
                 '3shot' : _3shot_idx,
                 '4shot' : _4shot_idx,
                 '5shot' : _5shot_idx,
                 '8shot' : _8shot_idx,
                 '16shot' : _16shot_idx,
                 '1pt' : _1pt_idx,
                 '10pt' : _10pt_idx}

for root, dirs, files in os.walk('../checkpoint'):
    for file in files:
        if ('max_results_channel_wise_epsilon.txt' == file) and ('naive_bsearch' == root.split('/')[-1]) and ('BSearch_recycle_adaptive_th_decouple_search_continue_channel_wise_new_no_cyan' in root.split('/')):
            path_list = root.split('/')
            dataset = path_list[3].split('_')[0]
            pretrain_method = path_list[2]
            arch = path_list[4]
            flag = path_list[5]
            if ('decouple' not in flag) or ('cri' not in path_list[6]):
                continue

            params = path_list[6].split('_')
            search_lr = params[7]
            if search_lr == '1.0':
                search_lr = '1'

            fc = params[11]
            criterion = params[15]
            th = params[17]
            M = params[21]
            right = params[19]
            if right == '10.0':
                right = '10'
            shot_or_pt = path_list[-2]
            pre_epochs = params[23]
            cycle_epochs = params[25]
            final_epochs = params[27]
            batch_size = params[9] + '(256)'

            rule_info = [(pre_epochs, pre_idx), (cycle_epochs, cycle_idx), (final_epochs, final_idx), (batch_size, batch_size_idx)]


            with open(os.path.join(root, file), 'r') as f:
                lines = f.readlines()

                line_list = lines[-1].strip().split(' ')
                lrs = line_list[5]
                naive_acc = line_list[13]
                bsearch_acc = line_list[21]

                naive_acc = acc_reformat(naive_acc)
                bsearch_acc = acc_reformat(bsearch_acc)

                epsilon = lines[-2].strip().split(' ')[3]

                try:
                    max_acc_check(lines, naive_acc, bsearch_acc)
                except:
                    print('fatal error : max_acc_check from {}'.format(os.path.join(root, file)))
                    sys.exit(-1)



            row_start = -1
            for i in range(1, ws.max_row+1):
                row = get_row(ws, i)
                location_principle_1 = all([gradient in row for gradient in [dataset, pretrain_method, arch, fc, criterion, M, right, search_lr, lrs, pre_epochs, cycle_epochs, final_epochs, batch_size]])
                location_principle_2 = check_location_charact(row, rule_info)
                if location_principle_1 and location_principle_2:
                    row_start = i
                    break


            if row_start != -1:
                for i in range(row_start, ws.max_row+2):
                    row = get_row(ws, i)
                    if row[th_idx] == '':
                        row_end = i
                        break


                for j in range(row_start, row_end):
                    row = get_row(ws, j)
                    if th == row[th_idx]:

                        ws.cell(j, shot_col_dict[shot_or_pt][0]+1).value = epsilon
                        ws.cell(j, shot_col_dict[shot_or_pt][1]+1).value = naive_acc
                        ws.cell(j, shot_col_dict[shot_or_pt][2]+1).value = bsearch_acc



wb.save('/mnt/data3/fumh/FSL/BSearch/RealFewShot_new_no_cyan/results/tmp_in_data3_latest.xlsx')